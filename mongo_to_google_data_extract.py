# -*- coding: utf-8 -*-
"""
Created on Tue Jul 30 13:55:19 2019

The pipeline to extract data from mongoDB and to save to google spread sheet
for annotation.

check_extraction.sav : The file that the hand-coded filtered data is stored.

@author: Babak Hosseini
"""
print(__doc__)

#from mongo_classifier import H_annot, class_code
from  mongo_classifier import H_annot
import pickle
import time
import os.path
import numpy as np
from datetime import datetime
import pandas as pd
import pymongo
from openpyxl import load_workbook
#from pymongo import Connection
from pymongo import MongoClient
import pprint
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import shutil
from git import Repo
from gspread_pandas import Spread

Mclient = MongoClient('mongodb://192.168.2.208:27018')

db = Mclient.sms
db.list_collection_names()

os.chdir('C:\\SYNX')

clients = db['client']
checks = db['check']
#%%==================== List of active devices for the month
# getting list of device_ids for work station and servers with specific api key!
pipelin3 = [
        
           {
           "$lookup":
                     {
                       "from": "site",
                       "localField": "siteid",
                       "foreignField": "_id",
                       "as": "site"
                      }    
            },
            {
                "$unwind": {"path": "$site", "preserveNullAndEmptyArrays": True}
            },
            {
                 "$lookup": 
                    {
                        "from": "client",
                        "localField": "site.clientid",
                        "foreignField": "_id",
                        "as": "client"
                    }      
            },
            {
                "$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}
            },
            {"$match": {"site.enabled" : True}},
            {"$match": {"client.apiKey":"ae0a4c75230afae756fcfecd3d2838cf"}},
#            {"$match": {"dscLocalDate": {"$gt":datetime(2019,7,1)}}},
#            {"$count": "device count"},
            {
                    "$project": {
                                    "device_name":"$name",
                                    "site_name":"$site.name",
                                    "client_name":"$client.name",
#                                    "dscLocalDate":"$dscLocalDate"
#                                    "name" : 1
                                }
            },            
         ] 
WK_list = list(db.workstation.aggregate(pipelin3))
SV_list = list(db.server.aggregate(pipelin3))
WK_db = pd.DataFrame(WK_list)
WK_db['Type'] = "workstation"
#WK_db = pd.DataFrame({values:WK_db.values, 'Type':'Workstation'}, index = index)
SV_db = pd.DataFrame(SV_list)
SV_db['Type'] = "server"
device_db = pd.concat([SV_db,WK_db], ignore_index = True)
device_db.head(2)
#%%===========================================================================
#==================== Extracting failed checks from scracth! - for the month
#===========================================================================
# %%  safety check!
modif_code = 'green'  # Enter this code if you really want to modify the spreadsheet
get_code = input('''
You are about to delete the "google spread sheet"
and the "check_extraction.sav" file
and rebuild them from scrath!! 
--------------------
You can also rename the target sheet and save file by "g_sheet_name" and "save_file"
if you want to start a different data extraction
However, you need to create the new google sheet online
and share it with the API key firs!
--------------------
If you know what you are doing, please enter the pass-code! ;-)
Pass-code:''')
if get_code != modif_code:
#    print('dsadsa')
    raise ValueError('''Wrong code! 
                Cannot continue!!''')

print('''If anything goes wrong, please check the google sheet and the git history 
      to get of the google sheet and check_DB before modification...''')


save_file = 'check_extraction_test.sav'
g_sheet_name = "Checks list_test"



#-----git push
if os.path.isfile(save_file):
    repo = Repo(os.getcwd())
    repo.index.add([save_file])
    repo.index.commit("backup of check_DB before modification")
    origin = repo.remotes.origin
    origin.push()


scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)

head_ind = 8

sheet_g = Spread(g_sheet_name)
sheet_g.open_sheet("Sheet1", create=True)
sheet_g.sheet.resize(rows=head_ind)


DB_col_list = ['device_name','Type','checkstatus',
                                    'description','servertime','last_fail',
                                    'client_name','site_name','extra',
                                    'dsc247','deviceid','checkid','consecutiveFails',
                                    'Label']
check_DB=pd.DataFrame(columns = DB_col_list)

## writing to excell file
#excel_path = 'check_list.xlsx'
#if os.path.isfile(excel_path):
#    os.remove(excel_path)

i = 0
#  %%===========   loop over the devices
while i < len(device_db):
    device_id = int(device_db['_id'][i])
    print('\nGetting checks for device_id:',i,'/',len(device_db),'(%s)' % (device_db['device_name'][i]),'...')    
    
    # here you change the datetime for start and the end of the extraction priod
    results = checks.find(
                {
                    "servertime":
                    {
                                    "$gte": datetime(2019,6,1,0,0,0),
                                    "$lte": datetime(2019,7,31,23,59,59)
                                    },    
                    "deviceid":device_id,
                    "checkstatus": {"$ne":"testok"},
                }
                ,projection={'datetime': False}
                )
    
    check_results = list(results)
    len_fails = len(check_results)
    print('number of failed checks:',len_fails)
#  %%
    if len(check_results) == 0:
        i+=1
        continue
    print('Prepaing the SQL table...')
    check_SQL=pd.DataFrame(check_results, columns = ['servertime','description',
                                                     'checkstatus','consecutiveFails','dsc247',
                                                     'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                    ['checkid','servertime'])#, ascending = False)                              
    temp = device_db.loc[i,['device_name','client_name','site_name','Type']]
    check_SQL[['device_name','client_name','site_name','Type']] = check_SQL.apply(lambda row: temp, axis = 1)      
    check_SQL = check_SQL.reset_index(drop = True)
    check_SQL['consFails'] = ''
    check_SQL['last_fail'] = ''
    check_SQL0 = check_SQL.copy()  # for debuging - todo remove
    
    check_current = check_SQL.loc[0,'checkid']
    check_SQL.loc[0,'consFails'] = check_SQL.loc[0,'consecutiveFails']
    check_SQL.loc[0,'last_fail'] = check_SQL.loc[0,'servertime']
    dsc247 = check_SQL['dsc247'][0]
   
    i_f = 1
 #  %%   loop over the check-fail rows
    while i_f < len(check_SQL):

        check_next = check_SQL.loc[i_f,'checkid']
        checkname = check_SQL.loc[i_f,'description']
        extra = check_SQL.loc[i_f,'extra']
        if H_annot(checkname,extra)=='ignore': # definite case of label 3 case
            check_SQL = check_SQL.drop(i_f).reset_index(drop = True)            
            continue
        seq = 0       
        if check_next == check_current:  # same check
            seq = 0  # flags the existing sequence

            a = check_SQL['last_fail'][i_f-1]
            b = check_SQL['servertime'][i_f]
            cons_b = check_SQL['consecutiveFails'][i_f]
            cons_a = check_SQL['consecutiveFails'][i_f-1]
            
            if ((b - a).total_seconds() < 3.5*3600): # 2-3hr consequative errors
                seq = 1
            elif (b.day-a.day == 1 or (b.day-a.day <0 and (b-a).total_seconds() < 24*3600) ): # next day sequence
                if dsc247 == 2 : # safety check or consecutiveFails
                    seq = 1
                else:    # look for in between testok!
                    
                    results = checks.find(
                                        {
                                            "servertime": {
                                                            "$gte": a,
                                                            "$lte": b
                                                            },    
                                            "deviceid":device_id,                                
                                            "checkstatus": "testok",
                                            "checkid": check_current                                            }
                                        )
                    temp_results = list(results)
                    if len(temp_results) == 0: # continues fail sequence
                        seq = 1
            if seq == 1:  # continues failing sequanece ==> clear it from the table    
                check_SQL.loc[i_f-1,'extra'] = check_SQL.loc[i_f,'extra']
                check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
                i_f -= 1
                
            check_SQL.loc[i_f,'last_fail'] = b
            check_SQL.loc[i_f,'consFails'] = cons_b
        else:  # it's a new check            
            # adding new row
            dsc247 = check_SQL['dsc247'][i_f]
            check_SQL.loc[i_f,'last_fail'] = check_SQL.loc[i_f,'servertime']
        
        if (check_next != check_current) | (seq == 0):
            # categorizing previous row#                
            checkname = check_SQL.loc[i_f-1,'description']
            extra = check_SQL.loc[i_f-1,'extra']
            pr = H_annot(checkname,extra)
            if pr == 'ignore':
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1
            elif pr != 'ND': #(pr=='nH')|(pr=='H'): # check-definite label H/N/Nan case
                temp = check_SQL.loc[i_f-1,DB_col_list[:len(DB_col_list)-1]]
                temp['consecutiveFails'] = check_SQL.loc[i_f-1,'consFails']
                temp['Label'] = pr
                check_DB= check_DB.append(temp,ignore_index=True)
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1            
                
        if check_SQL['consFails'][i_f]=='':
                check_SQL.loc[i_f,'consFails'] = check_SQL.loc[i_f,'consecutiveFails']
        check_current = check_next                
        i_f += 1

    # -------Annotate the last entry of SQL table
    i_f = len(check_SQL)-1
    checkname = check_SQL.loc[i_f,'description']
    extra = check_SQL.loc[i_f,'extra']
    pr = H_annot(checkname,extra)
    if pr == 'ignore': # check-definite label 3 case
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    elif pr != 'ND': #(pr=='nH')|(pr=='H'): # check-definite label H/N/Nan case
        temp = check_SQL.loc[i_f,DB_col_list[:len(DB_col_list)-1]]
        temp['consecutiveFails'] = check_SQL.loc[i_f,'consFails']
        temp['Label'] = pr
        check_DB= check_DB.append(temp,ignore_index=True)     
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    #  %%        
            
    check_SQL_last = check_SQL[['device_name','Type','checkstatus',
                                'description','servertime','last_fail',
                                'client_name','site_name','extra',
                                'dsc247','deviceid','checkid','consFails']]
    check_SQL_last = check_SQL_last.rename(columns = {"consFails":"consecutiveFails"})
    check_SQL_last = check_SQL_last.sort_values(by = 'servertime')
    
    #----------- save python data
    save_file = 'check_extraction.sav'
    pickle.dump({'device_i':i,'check_DB':check_DB}, open(save_file, 'wb'))
    
    #----------- save to google sheet
    if len(check_SQL_last) == 0:
        print('0 checks remained to save to the excel file...')
        i += 1
        continue
    
    print('Saving', len(check_SQL_last), 'extracted failes to the google sheet...')    
    sheet_g = Spread(g_sheet_name)
    sheet_g.open_sheet("Sheet1", create=False)  

    last_row = sheet_g.sheet.row_count
    sheet_g.df_to_sheet(check_SQL_last, index=False, headers = False,sheet='Sheet1', 
                            start=(last_row+2,1),  replace = False)

#---------saving to excell file
#    print('Saving', len(check_SQL_last), 'extracted failes to the excel file...')        
#    if not os.path.isfile(excel_path):
#        check_SQL_last.to_excel(excel_path, index = False)
#    else: # appending to the excell file
#        book = load_workbook(excel_path)
#        last_row = book['Sheet1'].max_row        
#        writer = pd.ExcelWriter(excel_path, engine='openpyxl')        
#        writer.book = book
#        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#        check_SQL_last.to_excel(writer, startrow = last_row+1,index = False, header = False)
#        writer.save()
    print("""
          =========== table saved =====
          =============================
          """
          )
    i += 1   
break    
#%%==================== If you want to try checks for a specific device or time priod 
results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2019,6,1,0,0,1),
                                    "$lte": datetime(2019,7,31,23,59,59),
#                                    "$ne": "false"
                                    },    
#                    "servertime": datetime(2019,3,1,18,51,21),
                    "deviceid":952849,
    #                "dsc247":2,
#                    "checkstatus": {"$ne":"testok"},  
#                    "description": 'Überprüfung des Dateisystemspeicherplatzes - Backup'
                    "checkid": "23573768"
                }
                ,projection={'datetime': False}
                )#.limit(50)
some_results = list(results)
partial_SQL=pd.DataFrame(some_results, columns = ['servertime','description',
                                                         'checkstatus','consecutiveFails','dsc247',
                                                         'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                        ['checkid','servertime'])#, ascending = False)    

partial_SQL.head(2)    