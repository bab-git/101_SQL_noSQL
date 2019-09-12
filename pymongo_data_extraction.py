# -*- coding: utf-8 -*-
"""
Created on Tue Jul 30 13:55:19 2019

The pipeline to extract data from mongoDB and saving in excell file
for further annotation.

A part is dedicated to data completion for case we need the trajectory of the fails on occation

@author: Babak Hosseini
"""
print(__doc__)

import pickle
import time
import os.path
import numpy as np
from datetime import datetime
import pandas as pd
import pymongo
from openpyxl import load_workbook
#from pymongo import Connection
from pymongo import MongoClient
import pprint
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import shutil
from git import Repo
from gspread_pandas import Spread

Mclient = MongoClient('mongodb://192.168.2.208:27018')

db = Mclient.sms
db.list_collection_names()

os.chdir('C:\\101_task')

clients = db['client']
checks = db['check']
#%%============
clients.count_documents({})
#%%============  read false sheet data
scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)

# Find a workbook by name and open the first sheet
# Make sure you use the right name here.
sheet = client.open("check-short-list").sheet1
#sheet = client.open("Checks comparison").sheet1
list_of_false = sheet.get_all_values()
false_SQL=pd.DataFrame(list_of_false)#, ascending = False)  
head_SQL = pd.Series(['name','H priority','extra','wrong','thresh_type','thresh'])
#head_SQL = false_SQL.iloc[1]
false_SQL = false_SQL[1:]
false_SQL = false_SQL.rename(columns = head_SQL)
wrong_checks = false_SQL.loc[(false_SQL['wrong'] == 'TRUE') | (false_SQL['H priority'] == 'NA'),'name']
normal_checks = false_SQL.loc[false_SQL['H priority'] == 'FALSE','name']
thresh_checks = false_SQL.loc[(false_SQL['H priority'] == 'TRUE') & (false_SQL['thresh_type'] != ''),'name']

def H_annot(checkname,extra):
    pr = 'ND'  #not determined
    #find wrong check
    wrong_find =list(filter(lambda i: checkname.find(i) >=0,wrong_checks))
    normal_find =list(filter(lambda i: checkname.find(i) >=0,normal_checks))
    thresh_find =list(filter(lambda i: checkname.find(i) >=0,thresh_checks))
    if wrong_find:
        pr = 'ignore'    
    elif normal_find:
        pr = 'nH'
    elif checkname.find('Terra Backup') >= 0:
        pr = 'H'
    elif thresh_find:
        ind_1 = extra.find('t:')
        if ind_1 >=0:
            tot_size = extra[ind_1+2:extra.find('GB')]
            tot_size=float(tot_size.replace('.','').replace(',','.'))
           
#            ind_F = 
            ext = extra[extra.find('Frei:')+5:]
            free_size = ext[:ext.find('GB')]
            free_size=float(free_size.replace('.','').replace(',','.'))
                        
            if free_size/tot_size < .2:  # 20% free threshold
                pr = 'H'  # high P   
            else:
                pr = 'nH'  # high P                   
#    elif checkname.find('Festplattenspeicherüberprüfung - Laufwerk') >= 0:
#        ind_1 = extra.find('t:')
#        if ind_1 >=0:
#            tot_size = extra[ind_1+2:extra.find('GB')]
#            tot_size=float(tot_size.replace('.','').replace(',','.'))
#           
##            ind_F = 
#            ext = extra[extra.find('Frei:')+5:]
#            free_size = ext[:ext.find('GB')]
#            free_size=float(free_size.replace('.','').replace(',','.'))
#                        
#            if free_size/tot_size < .2:  # 20% free threshold
#                pr = 'H'  # high P   
#            else:
#                pr = 'nH'  # high P   
    return pr
#%%================= test H_annot
#    H_annot()


#%%==================== List of active devices for the month
# getting list of device ids for WK ans servers
pipelin3 = [
        
           {
           "$lookup":
                     {
                       "from": "site",
                       "localField": "siteid",
                       "foreignField": "_id",
                       "as": "site"
                      }    
            },
            {
                "$unwind": {"path": "$site", "preserveNullAndEmptyArrays": True}
            },
            {
                 "$lookup": 
                    {
                        "from": "client",
                        "localField": "site.clientid",
                        "foreignField": "_id",
                        "as": "client"
                    }      
            },
            {
                "$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}
            },
            {"$match": {"site.enabled" : True}},
            {"$match": {"client.apiKey":"ae0a4c75230afae756fcfecd3d2838cf"}},
#            {"$match": {"dscLocalDate": {"$gt":datetime(2019,7,1)}}},
#            {"$count": "device count"},
            {
                    "$project": {
                                    "device_name":"$name",
                                    "site_name":"$site.name",
                                    "client_name":"$client.name",
#                                    "dscLocalDate":"$dscLocalDate"
#                                    "name" : 1
                                }
            },            
         ] 
WK_list = list(db.workstation.aggregate(pipelin3))
SV_list = list(db.server.aggregate(pipelin3))
WK_db = pd.DataFrame(WK_list)
WK_db['Type'] = "workstation"
#WK_db = pd.DataFrame({values:WK_db.values, 'Type':'Workstation'}, index = index)
SV_db = pd.DataFrame(SV_list)
SV_db['Type'] = "server"
device_db = pd.concat([SV_db,WK_db], ignore_index = True)
device_db.head(2)
#%%===========================================================================
#==================== Extracting faield checks from scracth! - for the month
#===========================================================================
save_file = 'check_extraction.sav'
#loaded_data = pickle.load( open(save_file, "rb" ))
#i = loaded_data ['device_i']
#check_DB = loaded_data['check_DB']

i = 0
year_prob=[]
DB_col_list = ['device_name','Type','checkstatus',
                                    'description','servertime','last_fail',
                                    'client_name','site_name','extra',
                                    'dsc247','deviceid','checkid','consecutiveFails',
                                    'Label']
check_DB=pd.DataFrame(columns = DB_col_list)
#check_DB=pd.DataFrame(columns = DB_col_list+['Label'])
excel_path = 'check_list.xlsx'
if os.path.isfile(excel_path):
    os.remove(excel_path)

#  %%===========   loop over the devices
while i < len(device_db):
#while i <= len(device_db):shab mi
#    i = 0
    #device_id = WK_list[i]['_id']    
    #  %%
#    device_id = int(device_db['_id'][i])
#    device_id=int(device_db['_id'][device_db['device_name']=='SRV-PR-01'])
    device_id = 625873
    i = device_db.loc[device_db['_id']==device_id,'_id'].index[0]            
    print('\nGetting checks for device_id:',i,'/',len(device_db),'(%s)' % (device_db['device_name'][i]),'...')    
#    del resultsd
    results = checks.find(
                {
                    "servertime":
                    {
                                    "$gte": datetime(2019,6,1,0,0,0),
                                    "$lte": datetime(2019,7,31,23,59,59)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":1054972,
    #                "dsc247":2, 
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "26706789"
                }
                ,projection={'datetime': False}
                )
#    try:
    check_results = list(results)
#    except pymongo.errors.InvalidBSON:
#       print('year format prblem ==> skip the device')
#       i+=1
#       year_prob.append(device_id)
#       break
    #  %%
    len_fails = len(check_results)
    print('number of failed checks:',len_fails)
#  %%
    if len(check_results) == 0:
        i+=1
        continue
    print('Prepaing the SQL table...')
    check_SQL=pd.DataFrame(check_results, columns = ['servertime','description',
                                                     'checkstatus','consecutiveFails','dsc247',
                                                     'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                    ['checkid','servertime'])#, ascending = False)                              
    temp = device_db.loc[i,['device_name','client_name','site_name','Type']]
    check_SQL[['device_name','client_name','site_name','Type']] = check_SQL.apply(lambda row: temp, axis = 1)      
    check_SQL = check_SQL.reset_index(drop = True)
    check_SQL['consFails'] = ''
    check_SQL['last_fail'] = ''
    check_SQL0 = check_SQL.copy()  # for debuging - todo remove
    
    check_current = check_SQL.loc[0,'checkid']
    check_SQL.loc[0,'consFails'] = check_SQL.loc[0,'consecutiveFails']
    check_SQL.loc[0,'last_fail'] = check_SQL.loc[0,'servertime']
    dsc247 = check_SQL['dsc247'][0]
   
    i_f = 1
 #  %%   loop over the check-fail rows
    while i_f < len(check_SQL):
#            if check_SQL['servertime'][i_f] >= datetime(2019,6,3,7,10,0):
#                break   
#  %%                 
        check_next = check_SQL.loc[i_f,'checkid']
#        if check_next == '26706789':
#            break        
        checkname = check_SQL.loc[i_f,'description']
        extra = check_SQL.loc[i_f,'extra']
        if H_annot(checkname,extra)=='ignore': # check-definite label 3 case
            check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
            #                print('ignore')
#            print('continue')
            continue
                
        if check_next == check_current:  # same check
            seq = 0  # flags the existing sequence

            a = check_SQL['last_fail'][i_f-1]
            b = check_SQL['servertime'][i_f]
            cons_b = check_SQL['consecutiveFails'][i_f]
            cons_a = check_SQL['consecutiveFails'][i_f-1]
            
            if ((b - a).total_seconds() < 3.5*3600): # 2-3hr consequative errors
                seq = 1
            elif (b.day-a.day == 1 or (b.day-a.day <0 and (b-a).total_seconds() < 24*3600) ): # next day sequence
                if dsc247 == 2 : # safety check or consecutiveFails
                    seq = 1
                else:    # look for in between testok!
                    
                    results = checks.find(
                                        {
                                            "servertime": {
                                                            "$gte": a,
                                                            "$lte": b
                                                            },    
                                            "deviceid":device_id,                                
                                            "checkstatus": "testok",
                                            "checkid": check_current                                            }
                                        )
                    temp_results = list(results)
                    if len(temp_results) == 0: # continues fail sequence
                        seq = 1
            if seq == 1:  # continues failing sequanece ==> clear it from the table    
                check_SQL.loc[i_f-1,'extra'] = check_SQL.loc[i_f,'extra']
                check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
                i_f -= 1
                
            check_SQL.loc[i_f,'last_fail'] = b
            check_SQL.loc[i_f,'consFails'] = cons_b
#            check_SQL.loc[i_f,'extra'] = extra
        else:  # new check            
            # adding new row
#                break
#                print('break')
            dsc247 = check_SQL['dsc247'][i_f]
            check_SQL.loc[i_f,'last_fail'] = check_SQL.loc[i_f,'servertime']
        
        if (check_next != check_current) | (seq == 0):
            # categorizing previous row#                
            checkname = check_SQL.loc[i_f-1,'description']
            extra = check_SQL.loc[i_f-1,'extra']
            pr = H_annot(checkname,extra)
            if pr == 'ignore':
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1
#                    print('ignore')
            elif (pr=='nH')|(pr=='H'): # check-definite label H/N case
#                    break
                temp = check_SQL.loc[i_f-1,DB_col_list[:len(DB_col_list)-1]]
                temp['consecutiveFails'] = check_SQL.loc[i_f-1,'consFails']
                temp['Label'] = pr
                check_DB= check_DB.append(temp,ignore_index=True)
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1
#                    print('Add')                    
#                    break                                        
                
        if check_SQL['consFails'][i_f]=='':
                check_SQL.loc[i_f,'consFails'] = check_SQL.loc[i_f,'consecutiveFails']
        check_current = check_next                
        i_f += 1
#  %%
    # -------Annotate last entry of SQL table
    i_f = len(check_SQL)-1
    checkname = check_SQL.loc[i_f,'description']
    extra = check_SQL.loc[i_f,'extra']
    pr = H_annot(checkname,extra)
    if pr == 'ignore': # check-definite label 3 case
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    elif (pr=='nH')|(pr=='H'): # check-definite label H/N case
        temp = check_SQL.loc[i_f,DB_col_list[:len(DB_col_list)-1]]
        temp['consecutiveFails'] = check_SQL.loc[i_f,'consFails']
        temp['Label'] = pr
        check_DB= check_DB.append(temp,ignore_index=True)     
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    #  %%        
            
    check_SQL_last = check_SQL[['device_name','Type','checkstatus',
                                'description','servertime','last_fail',
                                'client_name','site_name','extra',
                                'dsc247','deviceid','checkid','consFails']]
    check_SQL_last = check_SQL_last.rename(columns = {"consFails":"consecutiveFails"})
    check_SQL_last = check_SQL_last.sort_values(by = 'servertime')
    
    #----------- save python data
    save_file = 'check_extraction.sav'
    pickle.dump({'device_i':i,'check_DB':check_DB}, open(save_file, 'wb'))
    
    #----------- save to files
    if len(check_SQL_last) == 0:
        print('0 checks remained to save to the excel file...')
        i += 1
        continue
    
    print('Saving', len(check_SQL_last), 'extracted failes to the excel file...')        
    if not os.path.isfile(excel_path):
        check_SQL_last.to_excel(excel_path, index = False)
    else: # appending to the excell file
        book = load_workbook(excel_path)
        last_row = book['Sheet1'].max_row        
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')        
#        writer = pd.ExcelWriter(excel_path)
#        check_SQL_last.to_excel(writer)
#        writer.save()        
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        check_SQL_last.to_excel(writer, startrow = last_row+1,index = False, header = False)
        writer.save()
    print("""
          =========== tabel saved =====
          =============================
          """
          )
    i += 1    
#    loaded_data = pickle.load( open(filename, "rb" ))
#%%    
break  

#%%===========================================================================
#==================== Updating google sheet & checkDB checks for new device list
#===========================================================================
modif_code = 'green'  # Enter this code if you really want to modify the spreadsheet
get_code = input("DO YOU WANT TO MODIFY THE SPREAD SHEET?!\n")
if get_code != modif_code:
#    print('dsadsa')
    raise ValueError('You cannot modify the spread sheet!!')
print('Taking the backup of the google sheet and check_DB before modification...')
    

load_file = 'new_devices_list.sav'
loaded_data = pickle.load( open(load_file, "rb" ))
device_db_new = loaded_data['device_db_new']


load_file = 'check_extraction.sav'
loaded_data = pickle.load( open(load_file, "rb" ))
check_DB = loaded_data['check_DB']

head_ind=8        # index of the header
#label_col = 14

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)
sheet = client.open("Checks list").sheet1
sheet_g = Spread("Checks list")    
sheet_g.open_sheet("Sheet1", create=False)



#---- backup the data
headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
SQL_cpy = pd.DataFrame(all_values)
SQL_file = 'check_back_%s.xlsx' %(str(datetime.now().timestamp())[:10])
SQL_cpy.to_excel(SQL_file, index = False)


#-----git push
#dir_path = os.getcwd()
repo = Repo(os.getcwd())
#repo.heads.master.checkout()
#repo.index.add([DB_path])
repo.index.add([load_file])
repo.index.add([SQL_file])
repo.index.commit("backup of check_DB & google sheet before modification")
origin = repo.remotes.origin
origin.push()


#checks  = pd.DataFrame(all_values[head_ind:], columns = headers)
check_sheet  = pd.DataFrame(all_values, columns = headers)
#sheet_checks['servertime'][head_ind:] = [np.datetime64(a).astype(datetime) for a in sheet_checks['servertime'][head_ind:]]
#sheet_checks['last_fail'][head_ind:] = [np.datetime64(a).astype(datetime) for a in sheet_checks['last_fail'][head_ind:]]
check_list1 = (check_sheet['deviceid'][head_ind:])
check_list2 = (check_DB['deviceid'][head_ind:])
check_list = list(pd.concat([check_list1,check_list2]).unique())

#check_list = unicode check_list.

i = 0
#year_prob=[]
DB_col_list = ['device_name','Type','checkstatus',
                                    'description','servertime','last_fail',
                                    'client_name','site_name','extra',
                                    'dsc247','deviceid','checkid','consecutiveFails',
                                    'Label']
#excel_path = 'check_list.xlsx'
#if os.path.isfile(excel_path):
#    os.remove(excel_path)

# %%===========   loop over the devices
while i < len(device_db_new):
#while i <= len(device_db):shab mi
#    i = 0
    #device_id = WK_list[i]['_id']    
    #  %%
    device_id = int(device_db_new['_id'][i])
#    device_id = int(device_new[i])
#    device_id=int(device_db_new['_id'][device_db_new['device_name']=='SRV-PR-01'])
#    device_id = 625873
#    i = device_db_new.loc[device_db_new['_id']==device_id,'_id'].index[0]            
    if device_id in check_list:
        print(device_id, " device already added\n")
        i+=1
        continue

    print('\nGetting checks for device_id:',i,'/',len(device_db_new),'(%s)' % (device_db_new['device_name'][i]),'...')    
#    del resultsd
    results = checks.find(
                {
                    "servertime":
                    {
                                    "$gte": datetime(2019,6,1,0,0,0),
                                    "$lte": datetime(2019,7,31,23,59,59)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":1054972,
    #                "dsc247":2, 
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "26706789"
                }
                ,projection={'datetime': False}
                )
#    try:
    check_results = list(results)
#    except pymongo.errors.InvalidBSON:
#       print('year format prblem ==> skip the device')
#       i+=1
#       year_prob.append(device_id)
#       break
    #  %%
    len_fails = len(check_results)
    print('number of failed checks:',len_fails)
#  %%
    if len(check_results) == 0:
        i+=1
        continue
    print('Prepaing the SQL table...')
    check_SQL=pd.DataFrame(check_results, columns = ['servertime','description',
                                                     'checkstatus','consecutiveFails','dsc247',
                                                     'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                    ['checkid','servertime'])#, ascending = False)                              
    temp = device_db_new.loc[i,['device_name','client_name','site_name','Type']]
    check_SQL[['device_name','client_name','site_name','Type']] = check_SQL.apply(lambda row: temp, axis = 1)      
    check_SQL = check_SQL.reset_index(drop = True)
    check_SQL['consFails'] = ''
    check_SQL['last_fail'] = ''
    check_SQL0 = check_SQL.copy()  # for debuging - todo remove
    
    check_current = check_SQL.loc[0,'checkid']
    check_SQL.loc[0,'consFails'] = check_SQL.loc[0,'consecutiveFails']
    check_SQL.loc[0,'last_fail'] = check_SQL.loc[0,'servertime']
    dsc247 = check_SQL['dsc247'][0]
   
    i_f = 1
 #  %%   loop over the check-fail rows
    while i_f < len(check_SQL):
#            if check_SQL['servertime'][i_f] >= datetime(2019,6,3,7,10,0):
#                break   
#  %%                 
        check_next = check_SQL.loc[i_f,'checkid']
#        if check_next == '26706789':
#            break        
        checkname = check_SQL.loc[i_f,'description']
        extra = check_SQL.loc[i_f,'extra']
        if H_annot(checkname,extra)=='ignore': # check-definite label 3 case
            check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
            #                print('ignore')
#            print('continue')
            continue
                
        if check_next == check_current:  # same check
            seq = 0  # flags the existing sequence

            a = check_SQL['last_fail'][i_f-1]
            b = check_SQL['servertime'][i_f]
            cons_b = check_SQL['consecutiveFails'][i_f]
            cons_a = check_SQL['consecutiveFails'][i_f-1]
            
            if ((b - a).total_seconds() < 3.5*3600): # 2-3hr consequative errors
                seq = 1
            elif (b.day-a.day == 1 or (b.day-a.day <0 and (b-a).total_seconds() < 24*3600) ): # next day sequence
                if dsc247 == 2 : # safety check or consecutiveFails
                    seq = 1
                else:    # look for in between testok!
                    
                    results = checks.find(
                                        {
                                            "servertime": {
                                                            "$gte": a,
                                                            "$lte": b
                                                            },    
                                            "deviceid":device_id,                                
                                            "checkstatus": "testok",
                                            "checkid": check_current                                            }
                                        )
                    temp_results = list(results)
                    if len(temp_results) == 0: # continues fail sequence
                        seq = 1
            if seq == 1:  # continues failing sequanece ==> clear it from the table    
                check_SQL.loc[i_f-1,'extra'] = check_SQL.loc[i_f,'extra']
                check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
                i_f -= 1
                
            check_SQL.loc[i_f,'last_fail'] = b
            check_SQL.loc[i_f,'consFails'] = cons_b
            check_SQL.loc[i_f,'extra'] = extra
        else:  # new check            
            # adding new row
#                break
#                print('break')
            dsc247 = check_SQL['dsc247'][i_f]
            check_SQL.loc[i_f,'last_fail'] = check_SQL.loc[i_f,'servertime']
        
        if (check_next != check_current) | (seq == 0):
            # categorizing previous row#                
            checkname = check_SQL.loc[i_f-1,'description']
            extra = check_SQL.loc[i_f-1,'extra']
            pr = H_annot(checkname,extra)
            if pr == 'ignore':
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1
#                    print('ignore')
            elif (pr=='nH')|(pr=='H'): # check-definite label H/N case
#                    break
                temp = check_SQL.loc[i_f-1,DB_col_list[:len(DB_col_list)-1]]
                temp['consecutiveFails'] = check_SQL.loc[i_f-1,'consFails']
                temp['Label'] = pr
                check_DB= check_DB.append(temp,ignore_index=True)
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1
#                    print('Add')                    
#                    break                                        
                
        if check_SQL['consFails'][i_f]=='':
                check_SQL.loc[i_f,'consFails'] = check_SQL.loc[i_f,'consecutiveFails']
        check_current = check_next                
        i_f += 1
#  %%
    # -------Annotate last entry of SQL table
    i_f = len(check_SQL)-1
    checkname = check_SQL.loc[i_f,'description']
    extra = check_SQL.loc[i_f,'extra']
    pr = H_annot(checkname,extra)
    if pr == 'ignore': # check-definite label 3 case
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    elif (pr=='nH')|(pr=='H'): # check-definite label H/N case
        temp = check_SQL.loc[i_f,DB_col_list[:len(DB_col_list)-1]]
        temp['consecutiveFails'] = check_SQL.loc[i_f,'consFails']
        temp['Label'] = pr
        check_DB= check_DB.append(temp,ignore_index=True)     
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    #  %%        
            
    check_SQL_last = check_SQL[['device_name','Type','checkstatus',
                                'description','servertime','last_fail',
                                'client_name','site_name','extra',
                                'dsc247','deviceid','checkid','consFails']]
    check_SQL_last = check_SQL_last.rename(columns = {"consFails":"consecutiveFails"})
    check_SQL_last = check_SQL_last.sort_values(by = 'servertime')
    
    #----------- save python data
    save_file = 'check_extraction.sav'
    pickle.dump({'device_i':i,'check_DB':check_DB}, open(save_file, 'wb'))
    
    #----------- save to files
    if len(check_SQL_last) == 0:
        print('0 checks remained to save to the excel file...')
        i += 1
        continue
    
#    raise ValueError('update gsheet')
    
    print('Saving', len(check_SQL_last), 'extracted failes to the google sheet...')        

#    last_row = sheet.row_count
#    for i in range(0,len(check_SQL_last)+1):
#        sheet.insert_row(list(''), index = last_row+1) 
#    sheet = client.open("Checks list").sheet1
#    sheet.row_count
    
#    sheet_g = Spread("temp1")
    sheet_g = Spread("Checks list")
    sheet_g.open_sheet("Sheet1", create=False)    
    
    last_row = sheet_g.sheet.row_count        
    sheet_g.df_to_sheet(check_SQL_last, index=False, headers = False,sheet='Sheet1', 
                        start=(last_row+2,1),  replace = False)
#    sheet_g.sheet.resize(rows=sheet_g.sheet.row_count+1) 
    
    check_list.append(device_id)

    print("""
          =========== tabel saved =====
          =============================
          """
          )
    i += 1    
#    loaded_data = pickle.load( open(filename, "rb" ))





  
#%%===========================================================================
#==================== Updating g-sheet based on the short-list  (H-annot)
#===========================================================================    
modif_code = 'green'  # Enter this code if you really want to modify the spreadsheet
get_code = input("DO YOU WANT TO MODIFY THE SPREAD SHEET?!\n")
if get_code != modif_code:
#    print('dsadsa')
    raise ValueError('You cannot modify the spread sheet!!')
print('Taking the backup of the google sheet and check_DB before modification...')
    

load_file = 'check_extraction.sav'
loaded_data = pickle.load( open(load_file, "rb" ))
check_DB = loaded_data['check_DB']

head_ind=8        # index of the header
#label_col = 14

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)
sheet = client.open("Checks list").sheet1
#sheet2 = client.open("temp2").sheet1


cwd = os.getcwd()
shutil.copy(cwd+'\\'+load_file,cwd+'\\'+'check_extraction_back_%s.sav' %(str(datetime.now().timestamp())[:10]))

#fails_select = sheet.row_values(8)
headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
SQL_cpy = pd.DataFrame(all_values)
SQL_cpy.to_excel('check_back_%s.xlsx' %(str(datetime.now().timestamp())[:10]), index = False)
#print('check_modify_%s.xlsx' %(str(datetime.now())[:19]))

#checks  = pd.DataFrame(all_values[head_ind:], columns = headers)
sheet_checks  = pd.DataFrame(all_values, columns = headers)
#sheet_checks['servertime'][head_ind:] = [np.datetime64(a).astype(datetime) for a in sheet_checks['servertime'][head_ind:]]
#sheet_checks['last_fail'][head_ind:] = [np.datetime64(a).astype(datetime) for a in sheet_checks['last_fail'][head_ind:]]


i_check = 0
#  %%
while i_check< len(sheet_checks):

    g_rmd = 0
    DB_rmd = 0
    device_name = sheet_checks['device_name'][i_check]
    if device_name == '':
        i_check +=1
        continue
    checkname = sheet_checks['description'][i_check]
    extra = sheet_checks['extra'][i_check]
    pr = H_annot(checkname,extra)
    
    if pr == 'ignore': 
        print ('Row %d: Removing from google sheet' %(i_check+1))
        #        print (i_check)
#        raise ValueError('ignore')
#        continue    
    
    if (pr == 'H') | (pr == 'nH'):
        print ('Row %d: Transfering an %s label to checkDB' %(i_check+1,pr))
#        print (i_check)
#        raise ValueError('H or nH')
        
        g_row = sheet.row_values(i_check+1)
        SQL_row = sheet_checks.iloc[i_check].values.tolist()
        SQL_row [13:16]= list(filter(lambda x: x != '', SQL_row[13:16]))
        
        if g_row != SQL_row:
            raise ValueError('SQL is not syncronized with google sheet \n update the SQL table from google-sheet')
        
        DB_col_list = check_DB.columns;
        
        
        deviceid = int(sheet_checks.loc[i_check,'deviceid'])
#        deviceid = 971448
#        checkid = '30016492'
        checkid = sheet_checks.loc[i_check,'checkid']
        time =  datetime.strptime(sheet_checks.loc[i_check,'servertime'], '%Y-%m-%d %H:%M:%S')
#        time = datetime(2019,7,5,15,43,16)
        
        dev_ind = check_DB['deviceid'] == deviceid
        chk_ind = check_DB['checkid'] == checkid
        tm_ind = check_DB['servertime'] > time
        
        if len(check_DB.loc[dev_ind].index) == 0:
            ind_add = len(check_DB) # to the end of list
        elif len(check_DB.loc[dev_ind & chk_ind].index) == 0:
            ind_add = check_DB.loc[dev_ind].index[0]-1  # to the top of device list
        elif len(check_DB.loc[dev_ind & chk_ind & tm_ind ].index) == 0:
            ind_add = check_DB.loc[dev_ind & chk_ind].index[0]-1  # to the top of device-check list
        else:            
            ind_add = check_DB.loc[dev_ind & chk_ind & tm_ind ].index[0]-1  # after the device-check happened before this one

                    
        #--------- update check_DB                
        if ind_add == -1:
            ind_add =0
        check_DB1 = pd.DataFrame(check_DB.iloc[:ind_add+1])
        check_DB2 = pd.DataFrame(columns = DB_col_list)
        check_DB2.loc[0] = sheet_checks.loc[i_check,DB_col_list]
        check_DB2['Label'] = pr
        check_DB2['servertime'] = [np.datetime64(a).astype(datetime) for a in check_DB2['servertime']]
        check_DB2['last_fail'] = [np.datetime64(a).astype(datetime) for a in check_DB2['last_fail']]                
        check_DB2[['deviceid',
              'dsc247','consecutiveFails']] = check_DB2[['deviceid','dsc247',
                                          'consecutiveFails']].apply(pd.to_numeric, errors='coerce')
        
#        check_DB2['dsc247'] = [int(a) for a in check_DB2['dsc247']]
#        check_DB2['deviceid'] = [int(a) for a in check_DB2['deviceid']]
#        check_DB2['consecutiveFails'] = [int(a) for a in check_DB2['consecutiveFails']]                
        
        check_DB3 = pd.DataFrame(check_DB.iloc[ind_add+1:])
        check_DB = pd.concat([check_DB1, check_DB2, check_DB3],  ignore_index=True)        
        
    if (pr == 'ignore') | (pr == 'H') | (pr == 'nH'):
        #--------- update google sheet
        i_rmv = i_check
#        if (sheet_checks.loc[i_check+1,'deviceid'] == '') & (g_rmd == 0):
#            i_rmv = [i_check+1,i_check]
#            for row in i_rmv:
#                sheet.delete_row(row+1)
#            g_rmd = 1
        if g_rmd == 0:
            sheet.delete_row(i_rmv+1)
            g_rmd = 1
                
        #--------- update local sheet
        if DB_rmd  == 0:
            sheet_checks = sheet_checks.drop(i_rmv).reset_index(drop = True)        
            DB_rmd = 1
        
        
#        save_file = 'check_extraction.sav'
        pickle.dump({'check_DB':check_DB}, open(load_file, 'wb'))
#        g_rmd , DB_rmd = 0,0
        continue
    
    i_check +=1


#%%===========================================================================
#==================== Bringing not annotated cases or missed cases forward for annotation
#===========================================================================    
head_ind = 8

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)
sheet = client.open("Checks list").sheet1
sheet_g = Spread("Checks list")    
sheet_g.open_sheet("Sheet1", create=False)

headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
SQL_cpy = pd.DataFrame(all_values)
#SQL_cpy.to_excel('check_back_%s.xlsx' %(str(datetime.now().timestamp())[:10]), index = False)

check_SQL  = pd.DataFrame(all_values, columns = headers)    

label_list = sheet.col_values(headers.index('Label')+1)
labeled_rows = list(filter(lambda x: label_list[x] != '', range(head_ind,len(label_list))))
labeled_rows = [int(x) for x in labeled_rows]
annot_checks = check_SQL.loc[labeled_rows,'description'].unique()
all_checks = check_SQL.loc[range(head_ind,len(check_SQL)),'description'].unique()
rest_checks = check_SQL.loc[range(labeled_rows[len(labeled_rows)-1],len(check_SQL)),'description'].unique()
last_annot = labeled_rows[len(labeled_rows)-1]+1
    
if len(all_checks) > len(annot_checks):  # need to bring forward not annotated checks
    moved_dv = []
    for check_i in all_checks:        
        if check_i not in annot_checks: # new check
           
            row_ck = list(check_SQL['description']).index(check_i)
            row_dv = list(check_SQL['deviceid']).index(check_SQL.loc[row_ck,'deviceid'])
            deviceid = check_SQL.loc[row_ck,'deviceid']
            
            
            if (last_annot+1) >= row_dv: # right place: after last annotation
                moved_dv.append(temp.loc[temp.index[0],'deviceid'])
                continue
#            else:
#                cond1= check_SQL.loc[list(check_SQL['description']).index(check_i)-1,'description'] not in annot_checks 
#                cond2= check_SQL.loc[list(check_SQL['description']).index(check_i)-2,'description'] not in annot_checks            

#            if cond1 | cond2: # right place 
            elif deviceid in moved_dv:  # new check from already moved device
                continue   

            a = list(check_SQL.loc[range(row_dv),'deviceid'].unique())
            if a[len(a)-1] in moved_dv: # new check after another new check
                moved_dv.append(deviceid)
                continue                        
            else : # wrong place
                raise ValueError('new check')
                print('Bringing a not-annotated check forward - check:' , check_i)
                row_sql = list(check_SQL['description']).index(check_i)
                device_id = check_SQL.loc[row_sql,'deviceid']
                get_indexes = lambda x, xs: [i for (y, i) in zip(xs, range(len(xs))) if x == y]
                rows_ind = get_indexes(device_id,check_SQL['deviceid'])
                temp = check_SQL.iloc[rows_ind]
                
                rows_ind.reverse()
                for row in rows_ind:
                    sheet.delete_row(row+1)
                
                for i in range(0,len(temp)+1):
                    sheet.insert_row(list(''), index = last_annot+1)                
                sheet_g.df_to_sheet(temp, index=False, headers = False,sheet='Sheet1', start=(last_annot+2,1),  replace = False)
                
                all_values = sheet.get_all_values()
                check_SQL  = pd.DataFrame(all_values, columns = headers)
                label_list = sheet.col_values(headers.index('Label')+1)
                labeled_rows = list(filter(lambda x: label_list[x] != '', range(head_ind,len(label_list))))
                labeled_rows = [int(x) for x in labeled_rows]            
                last_annot = labeled_rows[len(labeled_rows)-1]+1
                
                moved_dv.append(temp.loc[temp.index[0],'deviceid'])

print('Samples of all new checks are brough forward')                

            
    
    
#%%==================== Check the running operations
#break    
a=db.current_op()
print(len(a['inprog'])-1,'operation is still running')
#if len(a['inprog']) > 1:
#    print("Still running operations")
#else:
#    print("All is good!")
        
#agg_expl = db.command('aggregate','check',pipeline=pipeline, explain=True)

#            {"$group": {"_id":"$deviceid", "max_time":{"$max":"$servertime"}}}    


#print(client.f)

#connection = Connection()
#connection = Connection()
#%%==================== Check specific cases
results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2019,6,1,0,0,1),
                                    "$lte": datetime(2019,7,31,23,59,59),
#                                    "$ne": "false"
                                    },    
#                    "servertime": datetime(2019,3,1,18,51,21),
                    "deviceid":952849,
    #                "dsc247":2,
#                    "checkstatus": {"$ne":"testok"},  
#                    "description": 'Überprüfung des Dateisystemspeicherplatzes - Backup'
                    "checkid": "23573768"
                }
                ,projection={'datetime': False}
                )#.limit(50)
some_results = list(results)
partial_SQL=pd.DataFrame(some_results, columns = ['servertime','description',
                                                         'checkstatus','consecutiveFails','dsc247',
                                                         'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                        ['checkid','servertime'])#, ascending = False)    

partial_SQL.head(2)
#ISODate("2019-03-01T16:39:48.000Z") 1141533
#2019-03-01 18:09:40.000Z 745934
#2019-03-01 16:39:48.000Z
# %% performance test
del results
start = time.time()
results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2018,2,1,0,0,0),
                                    "$lte": datetime(2018,12,1,1,0,0,0)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2,
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
                )
some_results = list(results)
end = time.time()
print(end - start)

del results
start = time.time()
for i_t in range(0,5):
#    print(i_t)
    results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2018,2+i_t*2,1,0,0,0),
                                    "$lte": datetime(2018,2+(i_t+1)*2,1,1,0,0,0)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2,
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
                )
    some_results = list(results)
end = time.time()
print(end - start)

del results
start = time.time()
for i_t in range(0,10):
#    print(i_t)
    results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2018,2+i_t,1,0,0,0),
                                    "$lte": datetime(2018,2+i_t+1,1,1,0,0,0)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2,
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
                )
    some_results = list(results)
end = time.time()
print(end - start)

#%%=======================    test commands
pprint.pprint(clients.find_one(
        {
#                servertime: {"$gt": new ISODate("2019-07-26 01:00:55.000Z")}
            "name":"101 - Busold Consulting"
        }))
#%==============================================
result = checks.find_one(
        {
#                servertime: {"$gt": new ISODate("2019-07-26 01:00:55.000Z")}
            "deviceid":1156225
        })

pprint.pprint(result)
#%===================================
results = checks.find(
        {
            "deviceid":1156225
        }
        ).limit(2)

for result in results:
    pprint.pprint(result)
    #%%============ find last entry
pipeline = [ 
            {"$match":{
                        "$and":[
                                  {"servertime":
                                      {"$gte": datetime(2019,7,31),"$lt": datetime(2019,8,1)}},
                                  {"deviceid":1035046}
                                ]
                     }},            

            {"$group":
                {
                  "_id":"$deviceid",    
                  "max_time":{"$max":"$servertime"},
#                  "count": {"$sum":1},
                }}
            #{"$limit":1},                                                                     
           ]

#{"deviceid":1035046}
#{"$gte": 'new ISODate("2019-07-31 00:00:08.000Z")'}},
agg_result = list(db.check.aggregate(pipeline))[0]
pprint.pprint(agg_result)
T_end = agg_result['max_time']
#%%==================== last-1 entry
pipeline2 = [ 
            {"$match":{
                        "$and":[
                                  {"servertime":
                                      {"$gte": datetime(2019,7,31),"$lt": T_end}},
                                  {"deviceid":1035046}
                                ]
                     }},            

            {"$group":
                {
                  "_id":"$deviceid",    
                  "max_time":{"$max":"$servertime"},
#                  "count": {"$sum":1},
                }}
            #{"$limit":1},                                                                     
           ]
agg_result = list(db.check.aggregate(pipeline2))[0]
T_end_1 = agg_result['max_time']
Ts=T_end.hour-T_end_1.hour   # sampling time
# =======================  new devices
new_list = list(filter(lambda x: device_db.loc[x,'_id'] not in device_db1['_id'].values, range(len(device_db))))
device_db_new = device_db.iloc[new_list]