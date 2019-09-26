# -*- coding: utf-8 -*-
"""
Created on Mon Aug 19 16:43:27 2019

Analysis of the annotated data from mongoDB containing fail checks

@author: Babak Hosseini
"""

print(__doc__)
#from mongo_classifier import H_annot, class_code
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook
#from sklearn.base
from sklearn.tree import DecisionTreeClassifier, plot_tree
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier, ExtraTreesClassifier
from datetime import datetime
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn import metrics
from matplotlib import pyplot as plt
from sklearn.multiclass import OneVsRestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF
from sklearn.neural_network import MLPClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
from sklearn.datasets import make_classification, make_moons, make_circles
from matplotlib.colors import ListedColormap
from sklearn.svm import SVC
from sklearn import tree   

import pickle
import gspread
from oauth2client.service_account import ServiceAccountCredentials 

os.chdir('C:\\101_task')

# %% importing annotated data from excell file
#from gspread_pandas import Spread
#Spread.sheet_to_df

head_ind=8        # index of the header
label_col = 14

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)

# Find a workbook by name and open the first sheet
sheet = client.open("Checks list").sheet1

#read the label column from the spreadsheet
label_list = sheet.col_values(label_col)

labeled_rows = list(filter(lambda x: label_list[x] != '', range(head_ind,len(label_list))))
labeled_rows = [int(x) for x in labeled_rows]

#fails_select = sheet.row_values(8)
headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
fails  = pd.DataFrame([all_values[i] for i in labeled_rows] , columns = headers)

    
#reading from excel
#book = load_workbook(excel_path)
#last_row = book['Sheet1'].max_row
#print("%d rows in the excel file" % last_row)
#fails = pd.read_excel(excel_path, header = 7)
#fails.fillna('', inplace=True)
#fails.dtypes
#fails['excel_row'] = fails.index+9
#fails = fails.drop(fails[fails['Label']==''].index).reset_index(drop = True)
#fails = fails.drop(fails[fails['device_name']==''].index).reset_index(drop = True)

# %% quantifying the features + combine with checkDB
load_file = 'check_extraction.sav'
loaded_data = pickle.load( open(load_file, "rb" ))
check_DB = loaded_data['check_DB']

#fails.drop(['description', 'servertime', 'device_name'])
col_select = ['checkstatus','consecutiveFails','dsc247',
                                              'checkid','deviceid','Label','servertime',
                                              'description','extra', 'last_fail',
                                              'Type']
fails_select1 = pd.DataFrame(fails, columns = col_select)

fails_select2 = pd.DataFrame(check_DB, columns = col_select)

#fails_select = pd.concat([fails_select1,fails_select2], ignore_index = True)
fails_select = fails_select1.copy()   # only not definite cases

check_drp = (fails_select['checkstatus']=='')|(fails_select['checkstatus']=='add')|(fails_select['Label']=='4')|(fails_select['Label']=='5')
fails_select = fails_select.drop(fails_select[check_drp].index).reset_index(drop = True)
#fails_select=fails_select.reset_index(drop = True)

#fails_select1['Label'] = fails_select1['Label'].apply(pd.to_numeric, errors='coerce')


# convert strings to numbers
check_stats = ['testerror','testalertdelayed','testcleared','test_inactive', 
             'testok_inactive','testerror_inactive','testok','add']
check_dic={}
for value in range(0,len(check_stats)):
    check_dic[check_stats[value]] = value+1

fails_select['checkstatus']=list(map(lambda x:check_dic[x],fails_select['checkstatus']))
fails_select['servertime'] = [np.datetime64(a).astype(datetime) for a in fails_select['servertime']]

#fails_select['servertime'] = list(map(lambda x: np.datetime64(x).astype(datetime),fails_select['servertime']))
#time_index = list(map(lambda x: np.datetime64(x).astype(datetime),fails_select['servertime']))
fails_select[['deviceid','checkid',
              'dsc247','consecutiveFails']] = fails_select[['deviceid','checkid',
                                          'dsc247','consecutiveFails']].apply(pd.to_numeric, errors='coerce')

# ----------------- quantify check description
check_names = fails_select.description.unique()

#ind=list(filter(lambda i: check_names[i].find('PING-Überprüfung')>=0, range(len(check_names))))
#check_names[ind] = 'PING-Überprüfung'
#check_names = np.unique(check_names)


check_list = {}
for value in range(0,len(check_names)):
    check_list[check_names[value]] = value+1
fails_select = fails_select.drop('checkid', axis = 1)
fails_select['checkid'] = [check_list[i] for i in fails_select['description']]
check_list_inv = {k:v for (v,k) in check_list.items()}

# only H and nH:
#fails_select.loc[(fails_select['Label']=='3') | (fails_select['Label']=='1'),'Label'] = 'nH'
#fails_select.loc[fails_select['Label']=='2','Label'] = 'H'

# %% to numpy conversion
#feat_names = ['checkstatus','consecutiveFails','dsc247','checkid','Label']
feat_names = ['checkstatus','consecutiveFails','dsc247','checkid']
fails_mat = fails_select[feat_names].to_numpy()
#np.c_[fails_mat,np.array(fails_select['servertime'])]
X = fails_mat[:,0:4].copy()
#X = np.delete(X, [range(3)] , axis = 1)
#X[X[:,1]>1,1] = 2


#y = fails_mat[:,4].copy()
fails_select['Label'] = [int(i) for i in fails_select['Label']]
y = np.array(list(fails_select['Label']))

#y = np.random.randint(1,3,len(X))

#y[y<4] = 1;
#y[y==4] = 2;

#class_names = {1:'Normal', 2: 'High', 3: 'ignore', 4:'on watch', 5:'nan'}
#class_names = {1:'Normal', 2:'on watch'}
#class_names = {'nH':'not High', 'H':'High'}
#class_names = {1:'not High', 2:'High'}
class_names = {1:'normal', 2:'High', 3:'ignore'}

datasets = (X,y)
#%% manual pre-processing

#testok_inactive
np.where(y[X[:,0]==5]==2)
ind_rmv = np.where((X[:,0]==5))[0] # testok_inactive data which should be cosidered as ok

#testcleard
len(X[X[:,0]==3])
np.where(y[X[:,0]==3]==1)
y[X[:,0]==3]=1  # label all testcleard cases as normal

X[X[:,0]==2,0] = 1   # checkstatus: testalertdelayed == testerror
X[X[:,0]==3,0] = 1   # checkstatus: testcleared == testerror

X[X[:,1]>1,1] = 2    # consfail>2 -->2

X = np.delete(X,0, axis = 1)   # rmove checkstatus column
#ind_rmv = []

# removing one-instance checks
#print('One-instance checks are removed from the database!!')
#for i in X:
#    if len

#todo: take care of hand_coded cases first!
#todo: either train a classifier for them or hand-code the decision boundaries
#todo: some should be unified based on the first part of the description and add the last part as another feature

#Ereignisprotokollüberprüfung - Veeam Backup

#!!! all the H cases at the moment!
hand_coded_list = ['PING-Überprüfung',
                   'Ereignisprotokollüberprüfung',
                   'Sicherungsüberprüfung'
                   ]
#hand_ind=[]
for id in hand_coded_list:
#    list(filter(lambda name: name.find(id) >=0,fails_select['description']))
    ind_temp = list(filter(lambda ind: fails_select['description'][ind].find(id) >=0,range(len(fails_select))))
    ind_rmv = np.concatenate((ind_rmv,ind_temp))
#    hand_ind.
#    [i for fails_select['description'][i].find(id) >=0 ]

X = np.delete(X,ind_rmv, axis = 0)
y = np.delete(y,ind_rmv)
# %% classificaiton
#X, y = datasets

# only H and nH:
#y[(y==3) | (y==1)] = 1 #'nH'
#y[y!=2] = 1 #only 'H' and 'nH'

#---------- knowledge based classification: pre-annot list
#HERE!!

#---------- model training

#X=StandardScaler().fit_transform(X)
#X[:,3] = StandardScaler().fit_transform(X[:,3].reshape(-1,1)).reshape(1,-1)
X_train, X_test, y_train, y_test = \
        train_test_split(X,y,test_size = .2)#, random_state = 42)
#print(X_train[:4])
#X_train, X_test= \
#        np.delete(X_train, 1 , axis = 1), np.delete(X_test, 1 , axis = 1)
#X_train[X_train[:,1]>1,1] = 2
#X_test[X_test[:,1]>1,1] = 2
        
        
names = ["Nearest Neighbors", "Linear SVM", "RBF SVM", "Gaussian Process",
         "Decision Tree", "Random Forest", "Neural Net", "AdaBoost",
         "Naive Bayes"
#         ,"QDA"
         ]

classifiers = [
        KNeighborsClassifier(3),
        SVC(kernel="linear", C=0.025),
        SVC(gamma = 2, C=1),
        GaussianProcessClassifier(1.0* RBF(1)),
        DecisionTreeClassifier(max_depth = 10),
        RandomForestClassifier(max_depth = 5, n_estimators= 10, max_features = 1),
        MLPClassifier(alpha=1, max_iter=1000),
        AdaBoostClassifier(),
        GaussianNB()
#        ,QuadraticDiscriminantAnalysis()
        ]

#h = .02  # step size in the mesh
#  %%    
i = 1;
print("Labels:", np.unique(y_test),', names: ',class_names)
#fails_select.columns[[int(x) for x in np.unique(y_test)]
#figure = plt.figure(figsize=(10,7))
#for name, clf in zip(names,classifiers):
#        ax = plt.subplot(1, len(classifiers) + 1, i)
name , clf = names[4], classifiers[4]
clf = DecisionTreeClassifier(random_state = 0, min_samples_leaf = 2)
#    clf = DecisionTreeClassifier(random_state = 0, min_samples_leaf = 10)
#    clf = RandomForestClassifier(n_estimators= 250,random_state = 0)
clf.fit(X_train,y_train)
score = clf.score(X_test, y_test)
#score = clf.score(X_train, y_train)
print(name,' score : ',score)

y_pred = clf.predict(X_test)
print(metrics.confusion_matrix(y_test,y_pred, labels = np.unique(y_test)))
conf_mat = metrics.confusion_matrix(y_test,y_pred, labels = np.unique(y_test))
miss_ind = np.where(y_pred!=y_test)[0]
miss_ch = [check_list_inv[X_test[i,-1]] for i in miss_ind]


for i in miss_ind:      
    if X_test[i,-1] not in X_train[:,-1]: # no training sample
#        name = list(filter(lambda item:item[1]== X_test[i,-1], check_list.items()))[0]
        name = check_list_inv[X_test[i,-1]]
        print('No training for data: %d , class: %s - check: %s, ' % (i,name,str(y_test[i])))
#            print ('Bad check! ind = %d , checkid = %d' % (i, X_test[i]), ' Train label:',y_train[X_train.flatten() == X_test[i]])

H_missed_trained = list(filter(lambda i: (X_test[i,-1] in X_train[:,-1]) & (y_test[i]== 'H'), miss_ind))
H_missed_trained_nm = [check_list_inv[X_test[i,-1]] for i in H_missed_trained]

if H_missed_trained_nm:
    H_missed_trained_samples = list(filter(lambda i: X_train[i,-1] == X_test[H_missed_trained[0],-1], range(len(X_train))))


#  %%categorize all training data
clf_all = DecisionTreeClassifier(random_state = 0, min_samples_leaf = 2)
clf_all.fit(X,y)
score_all = clf_all.score(X, y)
y_pred_all = clf_all.predict(X)
print(score_all)
print(metrics.confusion_matrix(y,y_pred_all, labels = np.unique(y)))
conf_mat = metrics.confusion_matrix(y,y_pred_all, labels = np.unique(y))
miss_ind = np.where(y_pred_all!=y)[0]
miss_ch = [check_list_inv[X[i,-1]] for i in miss_ind]
miss_labels = [y[i] for i in miss_ind]

#fig, ax = plt.figure(figsize=(20,10))
fig, ax = plt.subplots(figsize=(40,10))
#tree.plot_tree(clf, filled=True, feature_names = feat_names, ax = ax , class_names = True);
#tree.plot_tree(clf, filled=True, feature_names = feat_names, ax = ax , class_names = list(class_names.values()));
tree.plot_tree(clf, filled=True, feature_names = feat_names, ax = ax , class_names = ['1','2']);

#tree.plot_tree(clf)

#ax.plot(x,iris[column])
# %%============ detail analysis
j=3

miss_high = list(filter(lambda i: y[i]==2,miss_ind))
miss_high_names = [check_list_inv[X[i,-1]] for i in miss_high]
temp_SQL = fails_select.loc[fails_select['description']==miss_high_names[j]]

ind_j = X[:,-1]==X[miss_high[j],-1]
ind_m = X[miss_high,-1]==X[miss_high[j],-1]
X_missed = X[miss_high]
print('\nX_label= \n',X[ind_j,:])
print('\nX_label_missed= \n',X_missed[ind_m,:])
print(y[ind_j])
print(miss_high_names[j])
print(len(np.where(ind_j)[0]))

#ind_clrd = np.where(fails_select['checkstatus']==3)[0]
#for i in ind_clrd:
#    device_id = fails_select.loc[i,'deviceid']
#    check_id = fails_select.loc[i,'checkid']
#    ind_rest = np.where((fails_select['deviceid']==device_id) &(fails_select['checkid']==check_id))[0]
#    if len(ind_rest)>1:
#        raise ValueError('len(ind_rest)>1')

#name = 'PING-Überprüfung'
#name = 'Ereignisprotokollüberprüfung'
#name = 'Sicherungsüberprüfung'
#name = 'Sicherungsüberprüfung - Microsoft Windows 7 Backup'
#name = 'Integritätsüberprüfung für physische Festplatte'
name = 'Skriptüberprüfung'
ind = np.where([name in str for str in fails_select['description']])[0]
#fails_select.loc[ind,'Label'].unique()
temp_SQL = fails_select.loc[ind]
#fails_mat[ind]
row_SQL = fails_select.loc[ind[0]]

ind = np.where(fails_select['Label']==2)[0]
temp_SQL2= fails_select.loc[ind]


ind = np.where([name in str for str in check_DB['description']])[0]
#ind = np.where(check_DB['Label']=='H')[0]
temp_DB= check_DB.loc[ind]



# %% ============== extract data rows from mongoDB and classify the evaluation set
#------ safety measure
modif_code = 'green'  # Enter this code if you really want to modify the spreadsheet
get_code = input("DO YOU WANT TO MODIFY THE SPREAD SHEET?!\n")
if get_code != modif_code:
#    print('dsadsa')
    raise ValueError('You cannot modify the spread sheet!!')
print('Taking the backup of the google sheet and check_DB before modification...')

#def class_code(row_SQL):
#    pr = 'ND'  #not determined
#    
#    #PING-Überprüfung
#    if row_SQL['description'].find('PING-Überprüfung')>=0:
#        if row_SQL['deviceid'] in {590715,801799,1090258}:
#            pr = 'H'
#        elif row_SQL['deviceid'] in {754547,620692}:
#            pr = 'nH'
#    elif row_SQL['description'].find('Ereignisprotokollüberprüfung')>=0 & row_SQL['description'].find('Backup')>=0:
#        if row_SQL['extra'] == 'Ereignis nicht gefunden':
#            pr = 'H'
#        elif row_SQL['extra'].find('successfully')>=0:
#            pr = 'ignore'        
#            
#            
#    return pr

#------------------
#import mongo_classifier import H_annot, class_code, des_split, label_pred   # this step is necessary to update the classifier data
import mongo_classifier # this step is necessary to update the classifier data

head_ind = 8

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)
sheet = client.open("Checks list").sheet1

sheet_g = Spread("Check evaluation")
sheet_g.open_sheet("Sheet1", create=False)   
sheet_g.sheet.resize(rows=head_ind)

headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
check_SQL  = pd.DataFrame(all_values, columns = headers) 

all_checks = list(check_SQL.loc[range(head_ind,len(check_SQL)),'description'].unique())


#====== validation loop

DB_col_list = ['device_name','Type','checkstatus',
                                    'description','servertime','last_fail',
                                    'client_name','site_name','extra',
                                    'dsc247','deviceid','checkid','consecutiveFails',
                                    'Label']
i_dev = 0
#  %%
while i_dev < len(device_db):
    device_id = int(device_db['_id'][i_dev])
    print('\nGetting checks for device_id:',i_dev,'/',len(device_db),'(%s)' % (device_db['device_name'][i_dev]),'...') 
    results = checks.find(
                {
                    "servertime":
                    {
                                    "$gte": datetime(2019,8,1,0,0,0),
                                    "$lte": datetime(2019,8,31,23,59,59)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":1054972,
    #                "dsc247":2, 
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "26706789"
                }
                ,projection={'datetime': False}
                )
    check_results = list(results)
    
    len_fails = len(check_results)
    print('number of failed checks:',len_fails)
    if len(check_results) == 0:
        i_dev+=1
        continue
    print('Prepaing the SQL table...')
    check_SQL=pd.DataFrame(check_results, columns = ['servertime','description',
                                                     'checkstatus','consecutiveFails','dsc247',
                                                     'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                    ['checkid','servertime'])#, ascending = False)     
   
    temp = device_db.loc[i_dev,['device_name','client_name','site_name','Type']]
    check_SQL[['device_name','client_name','site_name','Type']] = check_SQL.apply(lambda row: temp, axis = 1)      
    check_SQL = check_SQL.reset_index(drop = True)
    check_SQL['consFails'] = ''
    check_SQL['last_fail'] = ''
    check_SQL['Label'] = ''
    check_SQL0 = check_SQL.copy()  # for debuging - todo remove

    check_current = check_SQL.loc[0,'checkid']
    check_SQL.loc[0,'consFails'] = check_SQL.loc[0,'consecutiveFails']
    check_SQL.loc[0,'last_fail'] = check_SQL.loc[0,'servertime']
    dsc247 = check_SQL['dsc247'][0]
   
    i_f = 1    
 #  %%   loop over the check-fail rows
    while i_f < len(check_SQL):        
        check_next = check_SQL.loc[i_f,'checkid']
        checkname = check_SQL.loc[i_f,'description']
        extra = check_SQL.loc[i_f,'extra']
        
#        if H_annot(checkname,extra) in {'ignore','nH'}: # check-definite label 1, 3 case
        pr1 = mongo_classifier.H_annot(checkname,extra)
        pr2 = mongo_classifier.class_code(check_SQL.loc[i_f])            
        if pr1 in {'ignore','nH'} or pr2 in {'ignore','nH'}: # check-definite label 1,3 case            
            check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
            #                print('ignore')
#            print('continue')
            continue
                     
        if check_next == check_current:  # same check
            seq = 0  # flags the existing sequence

            a = check_SQL['last_fail'][i_f-1]
            b = check_SQL['servertime'][i_f]
            cons_b = check_SQL['consecutiveFails'][i_f]
            cons_a = check_SQL['consecutiveFails'][i_f-1]
            
            if ((b - a).total_seconds() < 3.5*3600): # 2-3hr consequative errors
                seq = 1
            elif (b.day-a.day == 1 or (b.day-a.day <0 and (b-a).total_seconds() < 24*3600) ): # next day sequence
                if dsc247 == 2 : # safety check or consecutiveFails
                    seq = 1
                else:    # look for in between testok!
                    
                    results = checks.find(
                                        {
                                            "servertime": {
                                                            "$gte": a,
                                                            "$lte": b
                                                            },    
                                            "deviceid":device_id,                                
                                            "checkstatus": "testok",
                                            "checkid": check_current                                            }
                                        )
                    temp_results = list(results)
                    if len(temp_results) == 0: # continues fail sequence
                        seq = 1
            if seq == 1:  # continues failing sequanece ==> clear it from the table    
                check_SQL.loc[i_f-1,'extra'] = check_SQL.loc[i_f,'extra']
                check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
                i_f -= 1
                
            check_SQL.loc[i_f,'last_fail'] = b
            check_SQL.loc[i_f,'consFails'] = cons_b
#            check_SQL.loc[i_f,'extra'] = extra
        else:  # new check            
            # adding new row
#                break
#                print('break')
            dsc247 = check_SQL['dsc247'][i_f]
            check_SQL.loc[i_f,'last_fail'] = check_SQL.loc[i_f,'servertime']
        
        if (check_next != check_current) | (seq == 0):
#            raise ValueError('(check_next != check_current) | (seq == 0):')
            # categorizing previous row#                
            checkname = check_SQL.loc[i_f-1,'description']
            extra = check_SQL.loc[i_f-1,'extra']
                        
            pr1 = mongo_classifier.H_annot(checkname,extra)
            pr2 = mongo_classifier.class_code(check_SQL.loc[i_f])
            
            if pr1 in {'ignore','nH'} or pr2 in {'ignore','nH'}: # check-definite label 1,3 case
                check_SQL = check_SQL.drop(i_f-1).reset_index(drop = True)
                i_f -= 1
            else: # Nan or ND or H
#                raise ValueError('end i_f')
                if pr1 == 'ND' and pr2 == 'ND': # not check-definite label H/Nan case
                    if checkname in all_checks:
                        pr = 'nH'
                    else:
                        pr = 'new'
                else: # one of them is 'H' or 'Nan'
                    pr = list(filter(lambda pr: pr in {'H','Nan'}, [pr1,pr2]))[0]
                
                check_SQL.loc[i_f-1,'Label'] = pr
                
        if check_SQL['consFails'][i_f]=='':
                check_SQL.loc[i_f,'consFails'] = check_SQL.loc[i_f,'consecutiveFails']
        check_current = check_next                
        i_f += 1
#  %%
    # -------Annotate last entry of SQL table
#    raise ValueError('end i_f')
    i_f = len(check_SQL)-1
    checkname = check_SQL.loc[i_f,'description']
    extra = check_SQL.loc[i_f,'extra']
    
    pr1 = mongo_classifier.H_annot(checkname,extra)
    pr2 = mongo_classifier.class_code(check_SQL.loc[i_f])
        
    if pr1 in {'ignore','nH'} or pr2 in {'ignore','nH'}: # check-definite label 1,3 case
        check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
    else: # Nan or ND or H
#        raise ValueError('end i_f')
        if pr1 == 'ND' and pr2 == 'ND': # not check-definite label H/Nan case
            if checkname in all_checks:
                pr = 'nH'
            else:
                pr = 'new'
        else: # one of them is 'H' or 'Nan'
            pr = list(filter(lambda pr: pr in {'H','Nan'}, [pr1,pr2]))[0]
                
        check_SQL.loc[i_f,'Label'] = pr
    #  %%        
            
    check_SQL_last = check_SQL[['device_name','Type','checkstatus',
                                'description','servertime','last_fail',
                                'client_name','site_name','extra',
                                'dsc247','deviceid','checkid','consFails','Label']]
    check_SQL_last = check_SQL_last.rename(columns = {"consFails":"consecutiveFails"})
    check_SQL_last = check_SQL_last.sort_values(by = 'servertime')
    
    
    #----------- save to server
    if len(check_SQL_last) == 0:
        print('0 checks remained to save to the excel file...')
        i_dev += 1
        continue
    
#    raise ValueError('to save')
    print('Saving', len(check_SQL_last), 'extracted failes to the google sheet...')        
    sheet_g = Spread("Check evaluation")
    sheet_g.open_sheet("Sheet1", create=False)    
    
    last_row = sheet_g.sheet.row_count        
    sheet_g.df_to_sheet(check_SQL_last, index=False, headers = False,sheet='Sheet1', 
                        start=(last_row+2,1),  replace = False)
    
    print("""
          =========== tabel saved in google sheet =====
          =============================
          """
          )
    i_dev += 1  
# %% feature extraction part
save_file = 'trained_classifier.sav'
import mongo_classifier

head_ind = 8

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)
sheet = client.open("Checks list").sheet1
headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
fails  = pd.DataFrame(all_values, columns = headers) 
check_drp = (fails['checkstatus']=='')|(fails['checkstatus']=='add')|(fails['Label']=='4')|(fails['Label']=='5')|(fails['Label']=='')
fails = fails.drop(fails[check_drp].index).reset_index(drop = True)
fails = fails.drop(0).reset_index(drop = True)
all_checks = list(fails.loc[range(head_ind,len(fails)),'description'].unique())


feat_names = ['sub_desc_qnt','extra_qnt','Type','Label']
#feat_names = ['sub_desc_qnt','extra_qnt','deviceid','Type','Label']
#feat_names = ['sub_desc_qnt','extra_qnt','deviceid','consecutiveFails','Label']

#coded_classes = []
class_cols = ['feat_name','split_name','classifier','label','score','feat_names','detail']
coded_classes = pd.DataFrame(columns = class_cols)

split_cand = []
group_feat = []
#for feat in all_checks:
#    if feat.find(' - ')>=0:
#        split_cand.append(feat)
#        name = feat[:feat.find(' - ')]
#        if name not in group_feat:
#            group_feat.append(name)
#split_cand = list(filter(lambda feat: feat.find(' - ')>=0, all_checks))
#len(split_cand)
#len(group_feat)
   
#  %% --------- extraction     
#name = 'Skriptüberprüfung'
#feat_i = split_cand[28]
i_name = 0
#while i_name < len(group_feat):
while i_name < len(all_checks):
    feat = all_checks[i_name]
    split = 0
    if feat.find(' - ')>=0:
        split_cand.append(feat)
        name = feat[:feat.find(' - ')]
#        split = 1
        if name not in group_feat:
            group_feat.append(name)
        else:
#            raise ValueError('name in the group')
            i_name += 1
            continue
    else:
        name = feat    

#    name = group_feat[i_name]

    # data preprations
    ind1 = np.where([name in str for str in fails['description']])[0]
    temp_SQL = fails.loc[ind1]
    
#    ind2 = np.where([name in str for str in check_DB['description']])[0]
#    temp_DB= check_DB.loc[ind2]
    
    col = ['description','extra','deviceid','Type','servertime','last_fail',
                                    'dsc247','consecutiveFails',
                                    'checkstatus', 'Label']
    
    label_dic = {'H':2,'nH':1,'1':1,'2':2,'3':3,'5':5,'Nan':5}
    server_dic = {'server':1,'workstation':2}
    
    check_SQL1 = temp_SQL[col]
#    check_SQL2 = temp_DB[col]
#    check_SQL=pd.concat([check_SQL1,check_SQL2]).reset_index(drop = True)
    check_SQL = check_SQL1.copy()
    
    check_SQL['Label'] = [label_dic[a] for a in check_SQL['Label']]
    check_SQL['Type'] = [server_dic[i] for i in check_SQL['Type']]
    check_SQL[['deviceid','dsc247',
               'consecutiveFails']] = check_SQL[['deviceid','dsc247',
                                  'consecutiveFails']].apply(pd.to_numeric, errors='coerce')
    
    check_drp = (check_SQL['checkstatus']=='add')|(check_SQL['Label']>=4)
    check_SQL = check_SQL.drop(check_SQL[check_drp].index).reset_index(drop = True)
    
    # description extraction
    key = name
    
    check_SQL['sub_description'] = [mongo_classifier.des_split(x,key) for x in check_SQL['description']]
    sub_names = check_SQL.sub_description.unique()
    sub_list = {}
    for value in range(0,len(sub_names)):
        sub_list[sub_names[value]] = value+1
    #fails_select = fails_select.drop('checkid', axis = 1)
    check_SQL['sub_desc_qnt'] = [sub_list[i] for i in check_SQL['sub_description']]
        
    extra_names = check_SQL.extra.unique()
    ex_list = {}
    for value in range(0,len(extra_names)):
        ex_list[extra_names[value]] = value+1
    #fails_select = fails_select.drop('checkid', axis = 1)
    check_SQL['extra_qnt'] = [ex_list[i] for i in check_SQL['extra']]
    
    df = pd.DataFrame(check_SQL,columns=['description'])
    df = df.drop_duplicates()
    # eliminate single samples
    df2 = check_SQL.groupby(['description']).size()
    des_solo = list(df2.loc[df2==1].index)
    
    solo_ind = list(filter(lambda x:check_SQL.loc[x,'description'] in des_solo, range(len(check_SQL))))
        
    check_SQL = check_SQL.drop(solo_ind).reset_index(drop = True)
    
    if len(check_SQL['Label'].unique()) ==1:
        # solo-class
        # save split_name
#        raise ValueError('solo class')
        
        detail={}
        detail['sub_list'] = sub_list
        detail['ex_list'] = ex_list
        detail['des_solo'] = des_solo
        
        coded_classes.loc[len(coded_classes)] = ['',name,'',check_SQL.loc[0,'Label'],1,feat_names,detail]
        
        i_name += 1
        continue
    elif len(check_SQL['Label'].unique()) ==0:
#        raise ValueError('no data')
        detail={}
        detail['sub_list'] = sub_list
        detail['ex_list'] = ex_list
        detail['des_solo'] = des_solo
        
        coded_classes.loc[len(coded_classes)] = ['',name,'',[],0,feat_names,detail]
        
        i_name += 1
        continue


    
    fails_mat = check_SQL[feat_names].to_numpy()
    
    ind_label = len(feat_names)-1;
    X = fails_mat[:,:ind_label]
    y = fails_mat[:,ind_label]
    
    clf_all = DecisionTreeClassifier(random_state = 0, min_samples_leaf = 2)
    clf_all.fit(X, y)
    score = clf_all.score(X, y)
    
#    temp_coded = mongo_classifier.encoded_class()
#    temp_coded.split_name = name
#    temp_coded.classifier = clf_all
#    coded_classes.append(temp_coded)
    detail={}
    detail['sub_list'] = sub_list
    detail['ex_list'] = ex_list
    detail['des_solo'] = des_solo
        
    coded_classes.loc[len(coded_classes)] = ['',name,clf_all,check_SQL['Label'].unique(), score,feat_names,detail]
    
    i_name += 1

coded_classes['score'] = [round(a,3) for a in coded_classes['score']]

pickle.dump({'coded_classes':coded_classes}, open(save_file, 'wb'))

print ('Finished')

    
# %%
X_train, X_test, y_train, y_test = \
        train_test_split(X,y,test_size = .2)#, random_state = 42)
        
# -----------------classifier
clf = DecisionTreeClassifier(random_state = 0, min_samples_leaf = 2)
#clf = RandomForestClassifier(max_depth = 5, n_estimators= 10, max_features = 1),
clf_all = DecisionTreeClassifier(random_state = 0, min_samples_leaf = 2)


clf.fit(X_train,y_train)
score = clf.score(X_test, y_test)
y_pred = clf.predict(X_test)
print(score)
print(metrics.confusion_matrix(y_test,y_pred, labels = np.unique(y)))
miss_ind = np.where(y_pred!=y_test)[0]
X_test[miss_ind]


clf_all.fit(X,y)
score_all = clf_all.score(X, y)
y_pred_all = clf_all.predict(X)
print(score_all)
print(metrics.confusion_matrix(y,y_pred_all, labels = np.unique(y)))
conf_mat = metrics.confusion_matrix(y,y_pred_all, labels = np.unique(y))
miss_ind_all = np.where(y_pred_all!=y)[0]
#miss_ch = [check_list_inv[X[i,-1]] for i in miss_ind]
#miss_labels = [y[i] for i in miss_ind]

#fig, ax = plt.figure(figsize=(20,10))
fig, ax = plt.subplots(figsize=(40,10))
tree.plot_tree(clf_all, filled=True, feature_names = feat_names, ax = ax , class_names = ['1','2','3']);

# %%feature relevance
forest = RandomForestClassifier(n_estimators= 250, random_state = 0)

forest.fit(X,y)
importance = forest.feature_importances_

importance = forest.feature_importances_
std = np.std([tree.feature_importances_ for tree in forest.estimators_], axis = 0)

indices = np.argsort(importance)[::-1]

# Print the feature ranking
print("Feature ranking:")

for f in range(X.shape[1]):
    print("%d. feature %d (%f)" % (f+1,indices[f],importance[indices[f]]))

plt.figure()
plt.title("Feature importances")
plt.bar(range(X.shape[1]), importance[indices],
        color="r", yerr=std[indices], align="center")
plt.xticks(feat_names[0:4],indices)
#plt.xticks(range(X.shape[1]),indices)
plt.xlim([-1, X.shape[1]])
#plt.title('Extra_trees')
plt.title('Random Forest')
plt.show()

# %% ============== Label prediction based on data from google sheet 
#from mongo_classifier import H_annot, class_code, label_pred   # this step is necessary to update the classifier data

head_ind = 8        # index of the header
label_col = 15

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)

# Find a workbook by name and open the first sheet
sheet = client.open("Check evaluation").sheet1

#read the label column from the spreadsheet
#label_list = sheet.col_values(label_col)
#labeled_rows = list(filter(lambda x: label_list[x] != '', range(head_ind,len(label_list))))
#labeled_rows = [int(x) for x in labeled_rows]

check_list = sheet.col_values(11)
valid_rows = list(filter(lambda x: check_list[x] != '', range(head_ind,len(check_list))))
valid_rows = [int(x) for x in valid_rows]

#fails_select = sheet.row_values(8)
headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
SQL_test  = pd.DataFrame([all_values[i] for i in valid_rows] , columns = headers)
SQL_test0 =  SQL_test.copy()
SQL_test['pred Label'] = ''

# ====== classificartion loop
i_row = 0
while i_row < len(SQL_test):
    SQL_row = SQL_test.iloc[i_row]
    pr = mongo_classifier.label_pred(SQL_row)
    SQL_test.loc[i_row,'pred Label'] = pr
    i_row += 1
        
print ('all rows are analyzed')        
                            

#----------- saving the results    