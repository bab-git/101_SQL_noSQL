
"""
The pipeline extract sequene data for cases we need the trajectory of the fails 
to performs a temporal analysis on the data.

@author: Babak Hosseini
"""
print(__doc__)

import time
import os.path
import numpy as np
from datetime import datetime
import pandas as pd
import pymongo
from openpyxl import load_workbook
#from pymongo import Connection
from pymongo import MongoClient
import pprint
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_pandas import Spread

Mclient = MongoClient('mongodb://192.168.2.208:27018')

db = Mclient.sms
db.list_collection_names()

os.chdir('C:\\SYNX')

clients = db['client']
checks = db['check']
#%%==================== Extracting trajectory data on demand

modif_code = 'green'  # Enter this code if you really want to modify the spreadsheet
get_code = input('''
You are about to delete the existing "google spread sheet" and update it 
with new trajectories extracted from the "check list" sheet!! 
--------------------
You can also rename the target sheet and save file by "traj_sheet_name"
if you want to work on a different data set
However, you need to create the new google sheet online
and share it with the API key first!
--------------------
If you know what you are doing, please enter the pass-code! ;-)
Pass-code:''')
if get_code != modif_code:
#    print('dsadsa')
    raise ValueError('''Wrong code! 
                Cannot continue!!''')

print('''If anything goes wrong, please check the google sheet history 
      to get of the google sheet and check_DB before modification...''')



#start_ind = 8    # start and end row indexes to be processed in the excell file
#end_ind = 209
head_ind=8        # index of the header

scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('C:/Keys/mongoDB_secret.json', scope)
client = gspread.authorize(creds)

traj_sheet_name = "Check_trajectory"
sheet = client.open("Checks list").sheet1

#read the label column from the spreadsheet
label_list = sheet.col_values(14)

list_4 = list(filter(lambda x: label_list[x] == '4', range(len(label_list))))

#fails_select = sheet.row_values(8)
headers = sheet.row_values(head_ind)
all_values = sheet.get_all_values()
fails_select  = pd.DataFrame([all_values[a] for a in list_4], columns = headers)
fails_select['servertime'] = [np.datetime64(a).astype(datetime) for a in fails_select['servertime']]
fails_select['last_fail'] = [np.datetime64(a).astype(datetime) for a in fails_select['last_fail']]
fails_select['deviceid'] = [int(a) for a in fails_select['deviceid']]
fails_select['checkid'] = [str(int(a)) for a in fails_select['checkid']]

# %% Extraction loop
row_ind = 0
while row_ind < len(fails_select.index):
    start_time, end_time, checkid, deviceid = \
        fails_select.loc[row_ind,['servertime','last_fail','checkid', 'deviceid']]
    if end_time=='' : end_time = datetime(2019,7,31,23,59,59)
    deviceid = int(deviceid)
#, checkid = int(deviceid), str(int(checkid))
    print(('Getting checks for trajectoty: %d / %d') % (row_ind+1,len(fails_select.index)))
#    end_time = fails_select.loc[row_ind,'last_fail']
#    checkid = fails_select.loc[row_ind,'checkid']
#    deviceid = fails_select.loc[row_ind,'deviceid']
    results = checks.find(
                {
                    "servertime":
                    {
                                    "$gte": start_time,
                                    "$lte": end_time
                                    },    
                    "deviceid":deviceid,
                    "checkid": checkid
                }
                )    
    results = list(results)
#    len(check_results)
    check_SQL=pd.DataFrame(results, columns = ['checkstatus','description',
                                                         'servertime','extra','dsc247',
                                                         'deviceid','consecutiveFails','checkid']).sort_values(by = 
                                                        'servertime')#, ascending = False)
#                                                        ['checkid','servertime'])#, ascending = False)   
    temp = fails_select.loc[row_ind,['device_name','client_name','site_name','Type']]
    check_SQL[['device_name','client_name','site_name','Type']] = check_SQL.apply(lambda row: temp, axis = 1)      
    check_SQL_last = check_SQL[['device_name','Type','checkstatus',
                                    'description','servertime',
                                    'client_name','site_name','extra',
                                    'dsc247','deviceid','checkid','consecutiveFails']]
    
    print('Saving', len(check_SQL_last), 'extracted trajectory to the SQL table...')   

    sheet_T = Spread(traj_sheet_name)
    sheet_T.open_sheet("Sheet1", create=False)
    if row_ind == 0:
        sheet_T.df_to_sheet(check_SQL_last, index=False, sheet='Sheet1', replace = True)
    else:
        row = sheet_T.sheet.row_count+2
        sheet_T.df_to_sheet(check_SQL_last, index=False, headers = False,sheet='Sheet1', start=(row,1))

        
#------- save to excel file
#    excel_path = 'Check_trajectory.xlsx'
#    if not os.path.isfile(excel_path):
#        check_SQL_last.to_excel(excel_path, index = False)
#    else: # appending to the excell file
#        book = load_workbook(excel_path)
#        last_row = book['Sheet1'].max_row        
#        writer = pd.ExcelWriter(excel_path, engine='openpyxl')        
#        writer.book = book
#        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#        check_SQL_last.to_excel(writer, startrow = last_row+1,index = False, header = False)
#        writer.save()
    print("""
          =========== tabel saved =====
          ============================="""
          )
    row_ind +=1
print("""
      =========== All trajectories are saved in the excell file =====
      ============================================================="""
      )    
break
#%%==================== Check the running operations
a=db.current_op()
print(len(a['inprog'])-1,'operation is still running')