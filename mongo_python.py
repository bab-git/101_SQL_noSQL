# -*- coding: utf-8 -*-
"""
Created on Tue Jul 30 13:55:19 2019

@author: Babak Hosseini
"""
import time
import os.path
import numpy as np
from datetime import datetime
import pandas as pd
import pymongo
from openpyxl import load_workbook
#from pymongo import Connection
from pymongo import MongoClient
import pprint

Mclient = MongoClient('mongodb://192.168.2.208:27018')

db = Mclient.sms
db.list_collection_names()

os.chdir('C:\\101_task')

clients = db['client']
checks = db['check']
#%%=======================
pprint.pprint(clients.find_one(
        {
#                servertime: {"$gt": new ISODate("2019-07-26 01:00:55.000Z")}
            "name":"101 - Busold Consulting"
        }))
#%%=======================
result = checks.find_one(
        {
#                servertime: {"$gt": new ISODate("2019-07-26 01:00:55.000Z")}
            "deviceid":1156225
        })

pprint.pprint(result)
#%%============
results = checks.find(
        {
            "deviceid":1156225
        }
        ).limit(2)

for result in results:
    pprint.pprint(result)
#%%============
clients.count_documents({})
#%%============ ?? list of enabled devices

#%%============ find last entry
pipeline = [ 
            {"$match":{
                        "$and":[
                                  {"servertime":
                                      {"$gte": datetime(2019,7,31),"$lt": datetime(2019,8,1)}},
                                  {"deviceid":1035046}
                                ]
                     }},            

            {"$group":
                {
                  "_id":"$deviceid",    
                  "max_time":{"$max":"$servertime"},
#                  "count": {"$sum":1},
                }}
            #{"$limit":1},                                                                     
           ]

#{"deviceid":1035046}
#{"$gte": 'new ISODate("2019-07-31 00:00:08.000Z")'}},
agg_result = list(db.check.aggregate(pipeline))[0]
pprint.pprint(agg_result)
T_end = agg_result['max_time']
#%%==================== another sample
pipeline2 = [ 
            {"$match":{
                        "$and":[
                                  {"servertime":
                                      {"$gte": datetime(2019,7,31),"$lt": T_end}},
                                  {"deviceid":1035046}
                                ]
                     }},            

            {"$group":
                {
                  "_id":"$deviceid",    
                  "max_time":{"$max":"$servertime"},
#                  "count": {"$sum":1},
                }}
            #{"$limit":1},                                                                     
           ]
agg_result = list(db.check.aggregate(pipeline2))[0]
T_end_1 = agg_result['max_time']
Ts=T_end.hour-T_end_1.hour   # sampling time
#%%==================== List of active devices for the month
# getting list of device ids for WK ans servers
pipelin3 = [
        
           {
           "$lookup":
                     {
                       "from": "site",
                       "localField": "siteid",
                       "foreignField": "_id",
                       "as": "site"
                      }    
            },
            {
                "$unwind": {"path": "$site", "preserveNullAndEmptyArrays": True}
            },
            {
                 "$lookup": 
                    {
                        "from": "client",
                        "localField": "site.clientid",
                        "foreignField": "_id",
                        "as": "client"
                    }      
            },
            {
                "$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}
            },
            {"$match": {"site.enabled" : True}},
            {"$match": {"client.apiKey":"ae0a4c75230afae756fcfecd3d2838cf"}},
            {"$match": {"dscLocalDate": {"$gt":datetime(2019,7,1)}}},
#            {"$count": "device count"},
            {
                    "$project": {
                                    "device_name":"$name",
                                    "site_name":"$site.name",
                                    "client_name":"$client.name",
#                                    "name" : 1
                                }
            },            
         ] 
WK_list = list(db.workstation.aggregate(pipelin3))
SV_list = list(db.server.aggregate(pipelin3))
WK_db = pd.DataFrame(WK_list)
WK_db['Type'] = "workstation"
#WK_db = pd.DataFrame({values:WK_db.values, 'Type':'Workstation'}, index = index)
SV_db = pd.DataFrame(SV_list)
SV_db['Type'] = "server"
device_db = pd.concat([SV_db,WK_db], ignore_index = True)
device_db.head(2)
#%%==================== loop of getting faield checks - for the month
#i = 0
#year_prob=[]
while i <= len(device_db):
#while i <= len(device_db):shab mi
#    i = 0
    #device_id = WK_list[i]['_id']
    print('Getting checks for device_id:',i,'/',len(device_db),'...')
    device_id = int(device_db['_id'][i])
    #  %%
#    del resultsd
    results = checks.find(
                {
                    "servertime":
                    {
                                    "$gte": datetime(2019,7,30,0,0,0),
                                    "$lte": datetime(2019,7,31,23,59,59)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2, 
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
#                ,projection={ 'servertime': False }
                )
    try:
        check_results = list(results)
#    except Exception as ex:
    except pymongo.errors.InvalidBSON:
#       template = "An exception of type {0} occurred. Arguments:\n{1!r}"
#       message = template.format(type(ex).__name__, ex.args)
       print('year format prblem ==> skip the device')
       i+=1
       year_prob.append(device_id)
       continue
    #  %%
    len_fails = len(check_results)
    print('number of failed checks:',len_fails)
#    %%
    if len(check_results) > 0:
        print('Prepaing the SQL table...')
        check_SQL=pd.DataFrame(check_results, columns = ['servertime','description',
                                                         'checkstatus','consecutiveFails','dsc247',
                                                         'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                        ['checkid','servertime'])#, ascending = False)                              
        temp = device_db.loc[[i],['device_name','client_name','site_name','Type']].iloc[0]
        check_SQL[['device_name','client_name','site_name','Type']] = check_SQL.apply(lambda row: temp, axis = 1)      
        check_SQL = check_SQL.reset_index(drop = True)
        check_SQL['consFails'] = ''
        check_SQL['last_fail'] = ''
        check_SQL0 = check_SQL  # for debuging - todo remove
        
        check_current = check_SQL.loc[0,'checkid']
        check_SQL.loc[0,'consFails'] = check_SQL.loc[0,'consecutiveFails']
        dsc247 = check_SQL['dsc247'][0]
#        check_next = check_current
        
#        temp_SQL = check_SQL[check_SQL.index == 0]
#        ch_id_hist = np.array(check_SQL['checkid'][0])        
#        temp_SQL = pd.concat([temp_SQL,pd.DataFrame(columns = ['last_fail','index_last'])], sort=False)
#        check_SQL['last_fail'] = '' 
#        temp_SQL.loc[0,'last_fail'] = check_SQL['servertime'][0]
#        temp_SQL.loc[0,'index_last'] = 0
        i_f = 1
        i_g = 1
     #  %%   loop over the rows
        while i_f < len(check_SQL):
#            if check_SQL['servertime'][i_f] >= datetime(2019,6,3,7,10,0):
#                break                    
            check_next = check_SQL.loc[i_f,'checkid']
#            if check_next == '16880065':
#                break
            if check_next == check_current:  # same check
                seq = 0  # flags the existing sequence

#                i_match = temp_SQL.checkid == check_SQL['checkid'][i_f]
#                a = temp_SQL['last_fail'][i_match].reset_index(drop = True)[0]            
                if check_SQL['last_fail'][i_f-1]=='':
                    a = check_SQL['servertime'][i_f-1]
                else:
                    a = check_SQL['last_fail'][i_f-1]
                b = check_SQL['servertime'][i_f]
                cons_b = check_SQL['consecutiveFails'][i_f]
                cons_a = check_SQL['consecutiveFails'][i_f-1]
                
                if ((b - a).total_seconds() < 3.5*3600): # 2-3hr consequative errors
                    seq = 1
                elif (b.day-a.day == 1 or (b.day-a.day <0 and (b-a).total_seconds() < 24*3600) ): # next day sequence
                    if dsc247 == 2 : # safety check or consecutiveFails
                        seq = 1
                    else:    # look for in between testok!
                        
                        results = checks.find(
                                            {
                                                "servertime": {
                                                                "$gte": a,
                                                                "$lte": b
                                                                },    
                                                "deviceid":device_id,                                
                                                "checkstatus": "testok",
                                                "checkid": check_current                                            }
                                            )
                        temp_results = list(results)
                        if len(temp_results) == 0: # continues fail sequence
                            seq = 1
                if seq == 1:  # continues failing sequanece ==> clear it from the table    
                    check_SQL.loc[i_f-1,'extra'] = check_SQL.loc[i_f,'extra']
                    check_SQL = check_SQL.drop(i_f).reset_index(drop = True)
                    i_f -= 1
                    check_SQL.loc[i_f,'last_fail'] = b
                    check_SQL.loc[i_f,'consFails'] = cons_b                                        
            else:  # new check
                dsc247 = check_SQL['dsc247'][i_f]
                
#                check_current = check_next
#               ?? here?? dsc247 = check_SQL['dsc247'][i_f]
                
#                temp_SQL = temp_SQL.append(check_SQL.iloc[i_f],ignore_index=True)
#                i_match = len(temp_SQL)-1;        
#                ch_id_hist = np.append(ch_id_hist,check_SQL['checkid'][i_f])
            if check_SQL['consFails'][i_f]=='':
                    check_SQL.loc[i_f,'consFails'] = check_SQL.loc[i_f,'consecutiveFails']
            check_current = check_next                
#            temp_SQL.loc[i_match,'index_last'] = i_f        
#            temp_SQL.loc[i_match,'last_fail'] = b
            i_f += 1
            i_g += 1
        #  %%                        
        check_SQL_last = check_SQL[['device_name','Type','checkstatus',
                                    'description','servertime','last_fail',
                                    'client_name','site_name','extra',
                                    'dsc247','deviceid','checkid','consFails']]
        check_SQL_last = check_SQL_last.rename(columns = {"consFails":"consecutiveFails"})
        check_SQL_last = check_SQL_last.sort_values(by = 'servertime')
        
        excel_path = 'check_list.xlsx'
        print('Saving', len(check_SQL_last), 'extracted failes to the SQL table...')        
        if not os.path.isfile(excel_path):
            check_SQL_last.to_excel(excel_path, index = False)
        else: # appending to the excell file
            book = load_workbook(excel_path)
            last_row = book['Sheet1'].max_row        
            writer = pd.ExcelWriter(excel_path, engine='openpyxl')        
    #        writer = pd.ExcelWriter(excel_path)
    #        check_SQL_last.to_excel(writer)
    #        writer.save()        
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            check_SQL_last.to_excel(writer, startrow = last_row+1,index = False, header = False)
            writer.save()
        print("""
              =========== tabel saved =====
              ============================="""
              )
    i += 1
        
    
#%%
pd.DataFrame({values:df_her1.values,
                           'Percentage':price_ratio},
                            index = shorten_names).sort_values(by = 'preis',ascending=False)      
 

#%%==================== Check the running operations
a=db.current_op()
print(len(a['inprog'])-1,'operation is still running')
#if len(a['inprog']) > 1:
#    print("Still running operations")
#else:
#    print("All is good!")
        
#agg_expl = db.command('aggregate','check',pipeline=pipeline, explain=True)

#            {"$group": {"_id":"$deviceid", "max_time":{"$max":"$servertime"}}}    


#print(client.f)

#connection = Connection()
#connection = Connection()
#%%==================== Check specific cases
results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2019,6,1,0,0,1),
                                    "$lte": datetime(2019,7,31,23,59,59)
                                    },    
                    "deviceid":557108

,
    #                "dsc247":2,
#                    "checkstatus": {"$ne":"testok"},            
                    "checkid": "29965746"
                }
                )
some_results = list(results)
partial_SQL=pd.DataFrame(some_results, columns = ['servertime','description',
                                                         'checkstatus','consecutiveFails','dsc247',
                                                         'extra','checkid','deviceid']).sort_values(by = 
#                                                        'servertime')#, ascending = False)
                                                        ['checkid','servertime'])#, ascending = False)    

partial_SQL.head(2)
# %% performance test
del results
start = time.time()
results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2018,2,1,0,0,0),
                                    "$lte": datetime(2018,12,1,1,0,0,0)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2,
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
                )
some_results = list(results)
end = time.time()
print(end - start)

del results
start = time.time()
for i_t in range(0,5):
#    print(i_t)
    results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2018,2+i_t*2,1,0,0,0),
                                    "$lte": datetime(2018,2+(i_t+1)*2,1,1,0,0,0)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2,
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
                )
    some_results = list(results)
end = time.time()
print(end - start)

del results
start = time.time()
for i_t in range(0,10):
#    print(i_t)
    results = checks.find(
                {
                    "servertime": {
                                    "$gte": datetime(2018,2+i_t,1,0,0,0),
                                    "$lte": datetime(2018,2+i_t+1,1,1,0,0,0)
                                    },    
                    "deviceid":device_id,
#                    "deviceid":745976,
    #                "dsc247":2,
                    "checkstatus": {"$ne":"testok"},
#                    "checkid": "16880587"
                }
                )
    some_results = list(results)
end = time.time()
print(end - start)
#%% test
a=1
b=2
if (a == 1 and 
        b==2):
    print('dsadas')
